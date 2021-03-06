import datetime
import numpy as np
import operator

from math import ceil
from pprint import pprint as pp
from collections import Counter, OrderedDict

from .models import (Excel,
                     Sheet)

from django.views import View
from django.http import HttpResponse, JsonResponse
from openpyxl import load_workbook


class ExcelView(View):
    # 엑셀 파일 업로드
    def post(self, request):
        data = request.FILES.get('file', None)

        if data is None:
            return JsonResponse({"message": "DOES_NOT_FILE"}, status=400)

        # sheetList = []
        excel_name = str(data)

        try:

            if not data.name.endswith(".xlsx"):
                return JsonResponse({"message": "NOT_EXCEL_FILE"}, status=400)

            if Excel.objects.filter(name=excel_name).exists():
                return JsonResponse({"message": "EXISTS_EXCEL"}, status=400)

            Excel(
                name=excel_name
            ).save()

            wb = load_workbook(data, data_only=True)
            # [sheetList.append(i) for i in wb.sheetnames]
            sheetList = [i for i in wb.sheetnames]

            for sheet in sheetList:
                sheet_row = wb[sheet]
                all_values = []

                for row in sheet_row.rows:
                    row_value = [cell.value for cell in row]
                    all_values.append(row_value)

                for num, values in enumerate(all_values):
                    if not num == 0:
                        if values[15]:
                            Sheet.objects.create(
                                name=sheet,
                                Subset=values[1],
                                Compound_concentration_nM=values[2],
                                Replicate=values[3],
                                KaiChem_ID=values[4],
                                Compound_Name=values[5],
                                Compound_treatment_time=values[6],
                                Cell_line=values[7],
                                Plate_ID=values[8],
                                Well=values[9],
                                Sample_ID=values[10],
                                MGI_Index_No=values[11],
                                RNA_Extraction_date=values[12],
                                Library_Prep_date=values[13],
                                Sample_sending_date_LAS=values[14],
                                Sequencing_Completed=values[15],
                                RNA_quantity_ng=values[16],
                                DNA_quantity_ng=values[17],
                                excel_name_id=Excel.objects.get(
                                    name=excel_name).id
                            )
                        else:
                            Sheet.objects.create(
                                name=sheet,
                                Subset=values[1],
                                Compound_concentration_nM=values[2],
                                Replicate=values[3],
                                KaiChem_ID=values[4],
                                Compound_Name=values[5],
                                Compound_treatment_time=values[6],
                                Cell_line=values[7],
                                Plate_ID=values[8],
                                Well=values[9],
                                Sample_ID=values[10],
                                MGI_Index_No=values[11],
                                RNA_Extraction_date=values[12],
                                Library_Prep_date=values[13],
                                Sample_sending_date_LAS=values[14],
                                Sequencing_Completed='',
                                RNA_quantity_ng=values[16],
                                DNA_quantity_ng=values[17],
                                excel_name_id=Excel.objects.get(
                                    name=excel_name).id
                            )

            return HttpResponse(status=200)

        except KeyError:
            return JsonResponse({"message": "INVALID_KEY"}, status=400)

        except Exception as e:
            return JsonResponse({"message": e}, status=400)

    # excel 데이터 불러옴
    def get(self, request):
        try:
            excel_data = (Excel.
                          objects.
                          all().
                          values())

            excel_count = Excel.objects.count()

            return JsonResponse({"data": {
                "excel_data": list(excel_data),
                "excel_count": excel_count
            }}, status=200)

        except KeyError:
            return JsonResponse({"message", "INVALID_KEY"}, status=400)

        except TypeError:
            return JsonResponse({"message": "INVALID_TYPE"}, status=400)

        except Excel.DoesNotExist:
            return JsonResponse({"message": "DOESNOT_EXCEL"}, status=400)


class ExcelDetailView(View):
    # sheet 데이터를 불러오게 된다.
    def get(self, request, excel_name):

        if not Excel.objects.filter(name=excel_name).exists():
            return JsonResponse({"message": "DOESNOT_EXCEL"}, status=400)

        try:
            sheet_name = (Excel.
                          objects.
                          get(name=excel_name).
                          sheet_set.
                          values("name").
                          distinct())

            return JsonResponse({"sheet_data": list(sheet_name)}, status=200)

        except KeyError:
            return HttpResponse(status=400)

        except TypeError:
            return HttpResponse(status=400)

    # 엑셀 데이터를 삭제
    def delete(self, request, excel_name):

        if not Excel.objects.filter(name=excel_name).exists():
            return JsonResponse({"message", "DOESNOT_EXCEL"}, status=400)

        try:
            excel_name = Excel.objects.get(name=excel_name)
            excel_name.delete()

            return HttpResponse(status=200)

        except ValueError:
            return JsonResponse({"message", "INVALID_VALUE"}, status=400)

        except KeyError:
            return JsonResponse({"message", "INVALID_KEY"}, status=400)

        except Exception as e:
            return JsonResponse({"message", e}, status=400)


class SheetDetailView(View):
    # 엑셀 중에 특정 sheet 를 선택했을 경우
    def get(self, request, excel_name, sheet_name):

        if not Excel.objects.filter(name=excel_name).exists():
            return JsonResponse({"message": "DOESNOT_EXCEL"}, status=400)

        if not Sheet.objects.filter(name=sheet_name).exists():
            return JsonResponse({"message": "DOESNOT_SHEET"}, status=400)

        try:
            excel_id = Excel.objects.get(name=excel_name).id
            sheet_data = (Sheet.
                          objects.
                          filter(excel_name_id=excel_id,
                                 name=sheet_name).
                          values(
                              "Subset",
                              "Compound_concentration_nM",
                              "Replicate",
                              "KaiChem_ID",
                              "Compound_Name",
                              "Compound_treatment_time",
                              "Cell_line",
                              "Plate_ID",
                              "Well",
                              "Sample_ID",
                              "MGI_Index_No",
                              "RNA_Extraction_date",
                              "Library_Prep_date",
                              "Sample_sending_date_LAS",
                              "Sequencing_Completed",
                              "RNA_quantity_ng",
                              "DNA_quantity_ng"))

            cols_dict = []
            cols = ["id"] + [sheet for sheet in sheet_data[0]]
            cols_dict = [{"name": cols[num], "key":num}
                         for num in range(0, len(cols))]

            rows = []
            for num in range(0, len(sheet_data)):
                row = []
                [row.append(sheet) for sheet in sheet_data[num].values()]
                rows.append(row)

            return JsonResponse({"sheet_table": {
                "cols": cols_dict,
                "rows": rows,
            }}, status=200)

        except KeyError:
            return HttpResponse(status=400)

        except TypeError:
            return HttpResponse(status=400)

        except Exception as e:
            return JsonResponse({"message": e}, status=400)

# year , month , day 를 넘겨받았을때 weeks 로 바꿔준다.


def line_get_week_of_month(year, month, day):

    result = datetime.date(year, month, day).strftime("%V")
    if year == 2020:
        return int(result) % 40+1
    elif year == 2021:
        return int(result) + 14

# year , month , day 를 넘겨받았을때 weeks 로 바꿔준다.


def bar_get_week_of_month(year, month, day):
    dt = datetime.date(year, month, day)

    first_day = dt.replace(day=1)
    dom = dt.day
    adjusted_dom = dom + (1 + first_day.weekday()) % 7

    return int(ceil(adjusted_dom/7.0))


class StatisticsPage(View):
    def get(self, request):
        try:

            # circle
            kaichem_exclude = Sheet.objects.exclude(KaiChem_ID__in=["DMSO1",
                                                                    "DMSO2",
                                                                    "Niclo1",
                                                                    "Niclo2"]).values("KaiChem_ID").distinct().count()
            circle_number = kaichem_exclude * 100 // 1364

            # bar
            sheet = (Sheet.
                     objects.
                     exclude(KaiChem_ID__in=["DMSO1", "DMSO2", "Niclo1", "Niclo2"]))

            sequencing_data = [s.Sequencing_Completed for s in sheet]
            total_data_list = [[s.Library_Prep_date for s in sheet],
                               [d for d in sequencing_data if d != '' or None]]

            result_columns_data, columns_labels, bar_result, new_data_json = [
                [], []], [[], []], [[], []], [{}, {}]
            for tdl in range(0, len(total_data_list)):
                new_data_list = []

                for data in total_data_list[tdl]:
                    if data is None:
                        break

                    year, month, day = data[0:4], data[4:6], data[6:8]
                    new_data_list.append(
                        [year, month, bar_get_week_of_month(int(year), int(month), int(day))])

                for new_data in new_data_list:
                    if str(new_data[0]) not in new_data_json[tdl]:
                        new_data_json[tdl][str(new_data[0])] = dict()
                    if str(new_data[1]) not in new_data_json[tdl][str(new_data[0])]:
                        new_data_json[tdl][str(new_data[0])][str(new_data[1])] = [
                            0]*5
                    new_data_json[tdl][str(new_data[0])][str(
                        new_data[1])][new_data[2]-1] += 1

                ordered_d1 = dict(
                    **dict(OrderedDict(sorted(new_data_json[tdl].items()))))

                for year in ordered_d1:
                    for month in OrderedDict(sorted(new_data_json[tdl][year].items(), key=lambda t: t[0])):
                        for week in range(0, len(new_data_json[tdl][year][month])):
                            if new_data_json[tdl][year][month][week] != 0:
                                result_columns_data[tdl].append(
                                    new_data_json[tdl][year][month][week])
                                columns_labels[tdl].append(
                                    f"{year}{month}{week+1}")

            columns_labels_result = []

            for column in columns_labels:
                for c in column:
                    if len(c) != 0 and int(c) not in columns_labels_result:
                        columns_labels_result.append(int(c))

            for bar in range(0, len(bar_result)):
                for i in range(0, len(columns_labels_result)):
                    bar_result[bar].append(0)

            for index in range(0, len(new_data_json)):
                for year in new_data_json[index]:
                    for month in OrderedDict(sorted(new_data_json[index][year].items(), key=lambda t: t[0])):
                        for week in range(0, len(new_data_json[index][year][month])):
                            if new_data_json[index][str(year)][str(month)][week] > 0:
                                columns_index = columns_labels_result.index(
                                    int(str(year)+str(month) + str(week+1)))
                                bar_result[index][columns_index] = new_data_json[index][str(
                                    year)][str(month)][week]

            bar_labels_data = [
                f"{str(columns)[:4]}-{str(columns)[4:6]} {str(columns)[6:]}주" for columns in sorted(columns_labels_result)]

            while len(bar_labels_data) != 8:
                if len(bar_labels_data) < 8:
                    bar_labels_data.insert(0, "")
                elif len(bar_labels_data) > 8:
                    bar_labels_data = bar_labels_data[-8:]

            for bar_index in range(0, len(bar_result)):
                while len(bar_result[bar_index]) != 8:
                    if len(bar_result[bar_index]) < 8:
                        bar_result[bar_index].insert(0, 0)
                    elif len(bar_result[bar_index]) > 8:
                        bar_result[bar_index] = bar_result[bar_index][-8:]
            # line
            line_weeks_list = [i for i in range(1, 32)]
            line_last_result_list = [[], []]

            for tdl in range(0, len(total_data_list)):

                line_new_data_list = []
                line_new_data_json = dict()
                week_number = 0
                line_list = dict()

                for data in total_data_list[tdl]:
                    if data is None:
                        break

                    year, month, day = int(data[0:4]), int(
                        data[4:6]), int(data[6:8])
                    line_new_data_list.append(
                        [year, month, line_get_week_of_month(year, month, day)])

                for new_data in line_new_data_list:
                    if new_data[0] not in line_new_data_json:
                        line_new_data_json[new_data[0]] = dict()

                    if new_data[1] not in line_new_data_json[new_data[0]]:
                        line_new_data_json[new_data[0]][new_data[1]] = dict()

                    if new_data[2] not in line_new_data_json[new_data[0]][new_data[1]]:
                        line_new_data_json[new_data[0]
                                           ][new_data[1]][new_data[2]] = 0
                    line_new_data_json[new_data[0]
                                       ][new_data[1]][new_data[2]] += 1

                for year in OrderedDict(sorted(line_new_data_json.items(), key=lambda t: t[0])):
                    for month in OrderedDict(sorted(line_new_data_json[year].items(), key=lambda t: t[0])):
                        for week in OrderedDict(sorted(line_new_data_json[year][month].items(), key=lambda t: t[0])):
                            if week > week_number:
                                week_number = week
                            if line_new_data_json[year][month][week] > 0:
                                line_list[week] = line_new_data_json[year][month][week]

                week_number_result = [0 for i in range(1, week_number+1)]
                for line in line_list:
                    week_number_result[line-1] = line_list[line]

                line_count = 0

                for i in week_number_result:
                    line_count += i
                    line_last_result_list[tdl].append(line_count)

            return JsonResponse({
                "kaichem_number": kaichem_exclude,
                "circle_number": circle_number,
                "bar_labels": bar_labels_data,
                "bar_data": bar_result[0],
                "bar_data2": bar_result[1],
                "line_data": line_last_result_list[0],
                "line_data2": line_last_result_list[1],
                "line_weeks_list": line_weeks_list,
            }, status=200)

        except KeyError:
            return JsonResponse({"message": "INVALID_KEY"}, status=400)

        except Exception as e:
            return JsonResponse({"message": e}, status=400)
